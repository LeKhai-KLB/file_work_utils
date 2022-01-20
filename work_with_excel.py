from openpyxl import load_workbook, Workbook

def get_value_excel(file_path, sheetName, cellName):
    wb = load_workbook(filename = file_path)
    sheet = wb[sheetName]
    wb.close()
    return sheet[cellName].value
    
def update_value_excel(file_path, sheetName, cellName, value):
    wb = load_workbook(filename = file_path)
    sheet = wb[sheetName]
    sheet[cellName].value = value
    wb.close()
    wb.save(file_path)

def create_file_excel(file_path):
    wb = Workbook(file_path)
    wb.close()
    wb.save(file_path)

def add_new_sheet(file_path, sheet_name):
    wb = load_workbook(filename = file_path)
    wb.create_sheet(sheet_name)
    wb.close()
    wb.save(file_path)

def rename_sheet(file_path, sheet_name, new_name):
    wb = load_workbook(filename = file_path)
    wb[sheet_name].title = new_name
    wb.close()
    wb.save(file_path)

def get_all_sheet_names(file_path):
    wb = load_workbook(filename = file_path)
    sheets = wb.sheetnames
    wb.close()
    return sheets

if __name__ == '__main__':
    # filename = 'test.xlsx'
    # # result = get_value_excel(filename, 'A1')
    # # print (result)
    # update_value_excel(filename, 'Sheet 1', 'A2', '21')

    from pathlib import Path
    path = str(Path().cwd() / 'test.xlsx')

    create_file_excel(path)
    add_new_sheet(path, 'Data')
    rename_sheet(path, 'Data', 'New data')
    sheets = get_all_sheet_names(path)
